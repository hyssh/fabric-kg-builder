"""CSV / TSV / XLSX source loader — SPEC-002 §6.

Reads a tabular source file, produces:
- A ``SourceFileRow`` record (source_files table)
- A list of ``DocumentElementRow`` records (one per data row + one for the header)
- A ``schema-profile.json`` dict describing inferred column types and samples

Supported formats
-----------------
- ``.csv``   — comma-separated, BOM-safe, auto-detects delimiter
- ``.tsv``   — tab-separated
- ``.xlsx``  — each sheet treated as a separate logical table (requires openpyxl)

Error handling
--------------
- Malformed/empty CSV raises ``CsvLoaderError``
- Missing xlsx dependency raises ``CsvLoaderError`` with a helpful message
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fabric_kg_builder.model.ids import (
    content_hash as compute_content_hash,
    make_document_element_id,
    make_source_file_id,
)
from fabric_kg_builder.model.schemas import DocumentElementRow, SourceFileRow


# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------


class CsvLoaderError(ValueError):
    """Raised when CSV/TSV/XLSX loading fails due to malformed input or missing deps."""


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_SAMPLE_SIZE = 5  # values per column in schema-profile
_SUPPORTED_EXTENSIONS = {".csv", ".tsv", ".xlsx"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _file_content_hash(path: Path) -> str:
    """SHA-256 of raw file bytes."""
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    return digest


def _infer_type(values: list[str]) -> str:
    """Heuristic column type inference from a sample of non-null string values."""
    if not values:
        return "string"

    # Try integer
    int_count = sum(1 for v in values if v.strip().lstrip("-").isdigit())
    if int_count == len(values):
        return "integer"

    # Try float
    float_count = 0
    for v in values:
        try:
            float(v.strip())
            float_count += 1
        except ValueError:
            pass
    if float_count == len(values):
        return "float"

    return "string"


def _detect_delimiter(sample: str) -> str:
    """Use csv.Sniffer to detect delimiter; fall back to comma."""
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t|;")
        return dialect.delimiter
    except csv.Error:
        return ","


def _read_rows_from_text(
    text: str,
    delimiter: str | None = None,
) -> tuple[list[str], list[list[str]]]:
    """Parse CSV text into (headers, data_rows).

    Returns a tuple of (header list, list of row lists).
    Raises ``CsvLoaderError`` on empty or header-only files.
    """
    # Strip BOM
    text = text.lstrip("\ufeff")

    if not text.strip():
        raise CsvLoaderError("File is empty or contains only whitespace.")

    if delimiter is None:
        delimiter = _detect_delimiter(text[:4096])

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = [row for row in reader if any(cell.strip() for cell in row)]

    if not rows:
        raise CsvLoaderError("No data rows found after stripping blank lines.")

    headers = [h.strip() for h in rows[0]]
    if not headers:
        raise CsvLoaderError("Header row is empty.")

    data_rows = rows[1:]
    return headers, data_rows


def _build_column_profiles(
    headers: list[str], data_rows: list[list[str]]
) -> list[dict[str, Any]]:
    """Build schema-profile column entries for each header."""
    profiles: list[dict[str, Any]] = []
    for idx, name in enumerate(headers):
        values = [
            row[idx].strip()
            for row in data_rows
            if idx < len(row) and row[idx].strip()
        ]
        non_null = [v for v in values if v]
        null_count = len(data_rows) - len(non_null)
        unique_vals = list(dict.fromkeys(non_null))  # preserve order, deduplicate
        profiles.append(
            {
                "index": idx,
                "name": name,
                "inferred_type": _infer_type(non_null),
                "null_count": null_count,
                "unique_count": len(unique_vals),
                "sample_values": unique_vals[:_SAMPLE_SIZE],
            }
        )
    return profiles


def _row_content(headers: list[str], row: list[str]) -> str:
    """Serialize a CSV row as a compact key=value string for hashing/content."""
    pairs = []
    for h, v in zip(headers, row):
        pairs.append(f"{h}={v.strip()}")
    return "; ".join(pairs)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


class CsvLoadResult:
    """Result returned by :func:`load_csv`."""

    __slots__ = ("source_file", "document_elements", "schema_profile")

    def __init__(
        self,
        source_file: SourceFileRow,
        document_elements: list[DocumentElementRow],
        schema_profile: dict,
    ) -> None:
        self.source_file = source_file
        self.document_elements = document_elements
        self.schema_profile = schema_profile


def load_csv(
    path: str | Path,
    *,
    schema_profile_path: str | None = None,
    project_root: str | Path | None = None,
) -> CsvLoadResult:
    """Load a CSV, TSV, or XLSX file and return canonical records + schema profile.

    Parameters
    ----------
    path:
        Absolute or relative path to the source file.
    schema_profile_path:
        Optional path string to record in ``source_files.schema_profile_path``.
    project_root:
        Optional project root for computing the canonical relative path stored
        in ``source_files.path``.  Defaults to ``Path.cwd()``.

    Returns
    -------
    CsvLoadResult
        ``.source_file``        — ``SourceFileRow``
        ``.document_elements``  — list of ``DocumentElementRow`` (header + data rows)
        ``.schema_profile``     — dict matching SPEC-002 §6.2 schema

    Raises
    ------
    CsvLoaderError
        On empty files, malformed CSV, unsupported extension, or missing XLSX dep.
    FileNotFoundError
        When the file does not exist.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Source file not found: {path}")

    ext = path.suffix.lower()
    if ext not in _SUPPORTED_EXTENSIONS:
        raise CsvLoaderError(
            f"Unsupported file extension '{ext}'. "
            f"Supported: {', '.join(sorted(_SUPPORTED_EXTENSIONS))}"
        )

    if ext == ".xlsx":
        return _load_xlsx(path, schema_profile_path=schema_profile_path, project_root=project_root)

    return _load_delimited(
        path,
        delimiter="\t" if ext == ".tsv" else None,
        schema_profile_path=schema_profile_path,
        project_root=project_root,
    )


def _canonical_path(path: Path, project_root: Path | None) -> str:
    root = Path(project_root) if project_root else Path.cwd()
    try:
        rel = path.resolve().relative_to(root.resolve())
        return rel.as_posix()
    except ValueError:
        return path.as_posix()


def _load_delimited(
    path: Path,
    *,
    delimiter: str | None,
    schema_profile_path: str | None,
    project_root: Path | None,
) -> CsvLoadResult:
    """Internal: load CSV or TSV file."""
    raw_text = path.read_text(encoding="utf-8-sig")  # utf-8-sig strips BOM

    try:
        headers, data_rows = _read_rows_from_text(raw_text, delimiter=delimiter)
    except CsvLoaderError:
        raise
    except Exception as exc:  # pragma: no cover
        raise CsvLoaderError(f"Failed to parse {path.name}: {exc}") from exc

    return _build_result(
        path=path,
        headers=headers,
        data_rows=data_rows,
        source_type="tsv" if path.suffix.lower() == ".tsv" else "csv",
        schema_profile_path=schema_profile_path,
        project_root=project_root,
        sheet_name=None,
    )


def _load_xlsx(
    path: Path,
    *,
    schema_profile_path: str | None,
    project_root: Path | None,
) -> CsvLoadResult:
    """Internal: load first sheet of an XLSX file."""
    try:
        import openpyxl  # noqa: PLC0415
    except ImportError as exc:
        raise CsvLoaderError(
            "openpyxl is required to load XLSX files. "
            "Install it with: pip install openpyxl"
        ) from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    all_elements: list[DocumentElementRow] = []
    total_rows = 0

    # Process all sheets; build combined result using the first sheet as the
    # primary source_file row.
    first_result: CsvLoadResult | None = None

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        raw_rows = [[str(cell.value) if cell.value is not None else "" for cell in row] for row in ws.iter_rows()]
        # Filter blank rows
        raw_rows = [r for r in raw_rows if any(v.strip() for v in r)]
        if not raw_rows:
            continue
        headers = [h.strip() for h in raw_rows[0]]
        data_rows = raw_rows[1:]

        sheet_result = _build_result(
            path=path,
            headers=headers,
            data_rows=data_rows,
            source_type="xlsx",
            schema_profile_path=schema_profile_path,
            project_root=project_root,
            sheet_name=sheet_name,
        )
        total_rows += len(data_rows)
        all_elements.extend(sheet_result.document_elements)

        if first_result is None:
            first_result = sheet_result

    wb.close()

    if first_result is None:
        raise CsvLoaderError(f"XLSX file '{path.name}' contains no readable data sheets.")

    # Update row_count to span all sheets
    src = first_result.source_file.model_copy(update={"row_count": total_rows})
    return CsvLoadResult(
        source_file=src,
        document_elements=all_elements,
        schema_profile=first_result.schema_profile,
    )


def _build_result(
    path: Path,
    headers: list[str],
    data_rows: list[list[str]],
    source_type: str,
    schema_profile_path: str | None,
    project_root: Path | None,
    sheet_name: str | None,
) -> CsvLoadResult:
    now = datetime.now(timezone.utc)
    file_hash = _file_content_hash(path)
    can_path = _canonical_path(path, project_root)
    source_file_id = make_source_file_id(can_path, file_hash)

    source_file = SourceFileRow(
        source_file_id=source_file_id,
        path=can_path,
        filename=path.name,
        source_type=source_type,
        content_hash=file_hash,
        byte_size=path.stat().st_size,
        ingested_at=now,
        schema_profile_path=schema_profile_path,
        row_count=len(data_rows),
        notes=f"sheet={sheet_name}" if sheet_name else None,
    )

    document_elements: list[DocumentElementRow] = []

    # Header row element (element_type = "table")
    header_html = "<table><thead><tr>" + "".join(f"<th>{h}</th>" for h in headers) + "</tr></thead></table>"
    header_content = ", ".join(headers)
    header_hash = compute_content_hash(header_content)
    header_elem_id = make_document_element_id(source_file_id, "table", None, 0, header_hash)
    document_elements.append(
        DocumentElementRow(
            document_element_id=header_elem_id,
            source_file_id=source_file_id,
            element_type="table",
            title=sheet_name or path.stem,
            content=header_content,
            content_html=header_html,
            content_hash=header_hash,
            extracted_at=now,
            sort_order=0,
        )
    )

    # Data rows (element_type = "table_row")
    for i, row in enumerate(data_rows, start=1):
        row_content = _row_content(headers, row)
        row_hash = compute_content_hash(row_content)
        row_elem_id = make_document_element_id(source_file_id, "table_row", None, i, row_hash)
        document_elements.append(
            DocumentElementRow(
                document_element_id=row_elem_id,
                source_file_id=source_file_id,
                element_type="table_row",
                parent_element_id=header_elem_id,
                content=row_content,
                content_hash=row_hash,
                extracted_at=now,
                sort_order=i,
                row_index=i - 1,
            )
        )

    column_profiles = _build_column_profiles(headers, data_rows)
    schema_profile: dict[str, Any] = {
        "schema_profile_version": "1",
        "source_file_id": source_file_id,
        "source_path": can_path,
        "source_type": source_type,
        "inspected_at": now.strftime("%Y-%m-%dT%H:%M:%SZ"),
        "row_count": len(data_rows),
        "column_count": len(headers),
        "columns": column_profiles,
        "warnings": [],
    }

    return CsvLoadResult(
        source_file=source_file,
        document_elements=document_elements,
        schema_profile=schema_profile,
    )
